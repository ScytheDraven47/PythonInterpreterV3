# Ben Rogers-McKee (ScytheDraven47)
# 08/03/2017
#
# ExcelView for Interpreter Program

from openpyxl import Workbook, load_workbook

from views.file_view import FileView


class ExcelView(FileView):

    def __init__(self):
        self.has_loaded_excel = False
        self.has_created_excel = False
        self.has_received_data_from_excel = False
        self.has_filled_excel = False
        self.has_saved_excel = False

    def get_data(self, filename):
        """Gets a single row of data from the excel file, returns the data as an array"""
        wb = load_workbook(filename=filename)
        self.has_loaded_excel = True
        ws = wb['input']
        starting_col = 1
        all_data = []
        for r in range(2, ws.max_row):
            row_data = []
            for c in range(starting_col, ws.max_column + 1):
                row_data.append(ws.cell(row=r, column=c).value)
            all_data.append(row_data)
        self.has_received_data_from_excel = True
        return all_data

    def output(self, data_to_output, filename):
        """Loads the excel file, saves data(message) into the first empty row, saves file"""
        try:
            wb = load_workbook(filename)
            ws = wb['output']
            self.has_loaded_excel = True
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "output"
            self.has_created_excel = True
        row = 1
        while ws['A'+str(row)].value is not None:
            row += 1
        col = 1
        for row_data in data_to_output:
            for value in row_data.values():
                ws.cell(row=row, column=col).value = value
                col += 1
            row += 1
            self.has_filled_excel = True
        wb.save(filename)
        self.has_saved_excel = True
